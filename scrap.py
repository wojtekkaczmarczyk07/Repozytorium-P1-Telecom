import os
import time
import random
import logging
import traceback
import sqlite3
import re
import statistics
from collections import deque

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    StaleElementReferenceException,
    ElementClickInterceptedException,
    WebDriverException,
    NoSuchElementException,
)
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager

# ================== KONFIG ==================
EXCEL_INPUT = "numery.xlsx"
EXCEL_OUTPUT = "numery_z_operatorami.xlsx"
LOG_FILE = "app.log"

URL = "https://bip.uke.gov.pl/numeracja/dostawca-uslug/"
CACHE_DB = "operator_cache.db"

# Autotuning
TABS_INIT = 4               # startowa liczba aktywnych kart
MIN_TABS  = 2               # nigdy poniżej
MAX_TABS  = 6               # nigdy powyżej
AUTOTUNE_EVERY = 24         # co tylu wynikach przeliczamy i ewentualnie korygujemy liczbę kart

# Timingi i I/O
MAX_RETRIES = 4
RESULT_TIMEOUT = 8          # s – ile czekamy na wynik, zanim retry
BODY_CHANGE_TIMEOUT = 3     # s – ile czekamy aż body się zmieni po submit
POLL_INTERVAL = 0.06        # s – jak często „odpytywać” karty, gdy brak postępu
BASE_JITTER = (0.03, 0.08)  # podstawowy jitter między submitami (dodatkowo rośnie z liczbą kart)
SAVE_EVERY_RESULTS = 24     # zapis Excela po tylu NOWYCH wpisach
CACHE_COMMIT_EVERY = 20     # commit SQLite co tylu wpisach

# ================== LOGI ==================
logging.basicConfig(
    filename=LOG_FILE,
    filemode='a',
    encoding='utf-8',
    format='%(asctime)s %(levelname)s: %(message)s',
    level=logging.DEBUG
)
console = logging.StreamHandler()
console.setLevel(logging.INFO)
console.setFormatter(logging.Formatter('%(asctime)s %(levelname)s: %(message)s'))
logger = logging.getLogger()
logger.addHandler(console)

# ================== REGEXY ==================
RE_JEST = re.compile(r"operatorem\s+numeru.*?jest\s+([^()]+)", re.IGNORECASE | re.DOTALL)
RE_NALEZY = re.compile(r"należy\s+do\s+(.+)$", re.IGNORECASE)
RE_NIE_NALEZY = re.compile(r"nie\s+należy", re.IGNORECASE)

# ================== POMOCE ==================
def normalize_number(val):
    if val is None: return ''
    s = str(val).strip()
    return ''.join(ch for ch in s if ch.isdigit())

def get_driver():
    opts = ChromeOptions()
    opts.add_argument("--headless=new")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--window-size=1080,1500")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--log-level=3")
    opts.add_experimental_option("excludeSwitches", ["enable-logging", "enable-automation"])
    opts.page_load_strategy = "eager"

    prefs = {
        "profile.managed_default_content_settings.images": 2,
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False
    }
    opts.add_experimental_option("prefs", prefs)

    os.environ.setdefault("WDM_LOG_LEVEL", "0")
    service = ChromeService(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=opts)
    driver.set_page_load_timeout(20)

    try:
        driver.execute_cdp_cmd("Network.enable", {})
        driver.execute_cdp_cmd("Network.setBlockedURLs", {"urls": ["*.png", "*.jpg", "*.jpeg", "*.gif", "*.webp", "*.svg"]})
        driver.execute_cdp_cmd(
            "Page.addScriptToEvaluateOnNewDocument",
            {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"}
        )
    except Exception:
        pass
    return driver

def akceptuj_cookies(driver):
    try:
        btn = WebDriverWait(driver, 2).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(.,'Zgadzam')]"))
        )
        btn.click(); time.sleep(0.08); return
    except Exception:
        pass
    try:
        btn = driver.find_element(By.XPATH, "//span[@id='cmpbntyestxt']/parent::a")
        btn.click(); time.sleep(0.08); return
    except Exception:
        pass
    try:
        WebDriverWait(driver, 2).until(
            EC.frame_to_be_available_and_switch_to_it((By.CSS_SELECTOR, "iframe[src*='cookiebot']"))
        )
        btn = WebDriverWait(driver, 2).until(
            EC.element_to_be_clickable((By.XPATH, "//button[normalize-space(text())='Akceptuj wszystkie']"))
        )
        btn.click(); driver.switch_to.default_content(); time.sleep(0.08)
    except Exception:
        try: driver.switch_to.default_content()
        except Exception: pass

def ensure_search_page(driver, small_wait=4):
    WebDriverWait(driver, small_wait).until(EC.presence_of_element_located((By.ID, 'numer_telefonu')))
    WebDriverWait(driver, small_wait).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[type=submit].button')))

def robust_clear_input(driver, elem):
    try: driver.execute_script("arguments[0].value='';", elem)
    except Exception: pass
    try:
        elem.click(); elem.send_keys(Keys.CONTROL, "a"); elem.send_keys(Keys.BACK_SPACE)
    except Exception: pass

def wait_body_text_change(driver, before_text, timeout=BODY_CHANGE_TIMEOUT):
    end = time.time() + timeout
    while time.time() < end:
        try:
            now = driver.find_element(By.TAG_NAME, "body").text
            if now and now != before_text: return True
        except StaleElementReferenceException:
            return True
        except Exception:
            pass
        time.sleep(0.05)
    return False

def find_result_text(driver):
    patterns = ("operatorem numeru", "należy do", "nie należy")
    try:
        p = driver.find_element(
            By.XPATH,
            "//p[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'operatorem numeru')"
            " or contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'należy do')"
            " or contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'nie należy')]"
        )
        txt = p.text.strip()
        if any(k in txt.lower() for k in patterns):
            return txt
    except Exception:
        pass
    try:
        body = driver.find_element(By.TAG_NAME, "body").text
        low = body.lower()
        if any(k in low for k in patterns):
            for line in body.splitlines():
                l = line.strip()
                if any(k in l.lower() for k in patterns):
                    return l
            return body.strip()
    except Exception:
        pass
    raise TimeoutException("Brak tekstu wyniku.")

def submit_number(driver, numer):
    input_box = WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.ID, 'numer_telefonu')))
    submit_btn = WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type=submit].button")))
    robust_clear_input(driver, input_box)
    input_box.send_keys(numer)
    try:
        WebDriverWait(driver, 6).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[type=submit].button")))
        submit_btn.click()
    except (ElementClickInterceptedException, StaleElementReferenceException, WebDriverException):
        submit_btn = driver.find_element(By.CSS_SELECTOR, "input[type=submit].button")
        driver.execute_script("arguments[0].click();", submit_btn)

# ================== MULTI-TAB PIPELINE ==================
def open_tabs(driver, n):
    handles = [driver.current_window_handle]
    for _ in range(n - 1):
        driver.switch_to.new_window('tab')
        handles.append(driver.current_window_handle)
    return handles

def parse_operator(text):
    lower = text.lower()
    if RE_NIE_NALEZY.search(lower): return "Numer nieaktywny lub brak danych"
    m = RE_JEST.search(text)
    if m: return m.group(1).strip()
    m = RE_NALEZY.search(text)
    if m: return m.group(1).strip()
    return f"Nieznana odpowiedź: {text}"

def jitter_for_tabs(active_tabs):
    extra = max(0, active_tabs - 1) * 0.01
    a = BASE_JITTER[0] + extra
    b = BASE_JITTER[1] + extra
    return (a, b)

def median_safe(values, default=None):
    try:
        return statistics.median(values)
    except Exception:
        return default

def main():
    logger.info("Start (multi-tab pipeline + AUTOTUNE)")
    if not os.path.exists(EXCEL_INPUT):
        logger.critical(f"Brak pliku: {EXCEL_INPUT}")
        return

    # Excel: mapa numer -> wiersze
    wb = load_workbook(EXCEL_INPUT)
    sheet = wb.active
    if sheet.cell(1, 2).value is None:
        sheet.cell(1, 2).value = "Operator"

    num_to_rows = {}
    total_rows_pending = 0
    for idx in range(2, sheet.max_row + 1):
        numer = normalize_number(sheet.cell(idx, 1).value)
        if not numer: continue
        prev = str(sheet.cell(idx, 2).value or "").strip()
        if prev not in ["", "Nie znaleziono", None]: continue
        num_to_rows.setdefault(numer, []).append(idx)
        total_rows_pending += 1

    # Cache
    conn = sqlite3.connect(CACHE_DB)
    cur = conn.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS cache (numer TEXT PRIMARY KEY, operator TEXT)")
    conn.commit()

    cached = dict(cur.execute("SELECT numer, operator FROM cache").fetchall())

    to_fetch = []
    filled_from_cache = 0
    for numer, rows in num_to_rows.items():
        op = cached.get(numer)
        if op and not op.startswith("Błąd") and not op.startswith("Nieznana odpowiedź"):
            for r in rows: sheet.cell(r, 2).value = op
            filled_from_cache += len(rows)
        else:
            to_fetch.append(numer)

    logger.info(f"Do uzupełnienia: {total_rows_pending} (z cache: {filled_from_cache}). Zapytania online (unikalne): {len(to_fetch)}.")

    if filled_from_cache:
        try: wb.save(EXCEL_OUTPUT)
        except Exception as e: logger.warning(f"Zapis pośredni (cache) nieudany: {e}")

    if not to_fetch:
        logger.info("Wszystko z cache – kończę.")
        conn.close()
        try: wb.save(EXCEL_OUTPUT)
        except Exception as e: logger.error(f"Zapis końcowy nieudany: {e}")
        return

    # 1×Chrome, maksymalna liczba kart (niekoniecznie wszystkie aktywne)
    driver = get_driver()
    ALL_HANDLES = open_tabs(driver, min(MAX_TABS, max(1, len(to_fetch))))
    for h in ALL_HANDLES:
        driver.switch_to.window(h)
        driver.get(URL)
        akceptuj_cookies(driver)
        ensure_search_page(driver)

    # Stan kart
    tabs = {h: {'numer': None, 'attempt': 0, 'start_ts': 0.0, 'deadline': 0.0, 'latency': None} for h in ALL_HANDLES}

    # Kolejka numerów do zrobienia
    from collections import deque
    q = deque(to_fetch)

    # Autotuning: statystyki
    active_tabs = max(MIN_TABS, min(TABS_INIT, len(ALL_HANDLES)))
    last_tune_time = time.time()
    window_success = 0
    window_fail = 0
    window_timeouts = 0
    latencies_window = deque(maxlen=60)
    last_window_tps = None
    last_window_median = None

    # I/O liczniki
    processed_unique = 0
    wrote_since_save = 0
    pending_commits = 0
    total_unique = len(to_fetch)

    t0 = time.time()

    # Na start – zasiej zadania do „aktywnych” kart
    for h in ALL_HANDLES[:active_tabs]:
        if not q: break
        driver.switch_to.window(h)
        numer = q.popleft()
        try:
            try:
                body_before = driver.find_element(By.TAG_NAME, "body").text
            except Exception:
                body_before = ""
            submit_number(driver, numer)
            _ = wait_body_text_change(driver, body_before, timeout=BODY_CHANGE_TIMEOUT)
            tabs[h] = {'numer': numer, 'attempt': 1, 'start_ts': time.time(), 'deadline': time.time() + RESULT_TIMEOUT, 'latency': None}
            time.sleep(random.uniform(*jitter_for_tabs(active_tabs)))
        except Exception:
            # odśwież i spróbuj dalej w kolejnych iteracjach
            try: driver.refresh(); ensure_search_page(driver)
            except Exception: pass
            tabs[h] = {'numer': numer, 'attempt': 1, 'start_ts': time.time(), 'deadline': time.time() + RESULT_TIMEOUT, 'latency': None}

    try:
        while processed_unique < total_unique:
            progress = False
            active_set = set(ALL_HANDLES[:active_tabs])

            for h in ALL_HANDLES:
                st = tabs[h]
                numer = st['numer']

                # 1) Karta wolna – przydziel tylko jeśli jest „aktywna” wg autotuningu
                if not numer:
                    if (h in active_set) and q:
                        driver.switch_to.window(h)
                        try:
                            try:
                                body_before = driver.find_element(By.TAG_NAME, "body").text
                            except Exception:
                                body_before = ""
                            peek = q[0]
                            submit_number(driver, peek)
                            _ = wait_body_text_change(driver, body_before, timeout=BODY_CHANGE_TIMEOUT)
                            numer_new = q.popleft()
                            st.update({'numer': numer_new, 'attempt': 1, 'start_ts': time.time(),
                                       'deadline': time.time() + RESULT_TIMEOUT, 'latency': None})
                            time.sleep(random.uniform(*jitter_for_tabs(active_tabs)))
                            progress = True
                        except Exception:
                            try:
                                driver.refresh(); ensure_search_page(driver)
                            except Exception: pass
                    continue

                # 2) Czy przekroczyliśmy deadline? (retry lub fail)
                now = time.time()
                if now >= st['deadline']:
                    if st['attempt'] < MAX_RETRIES:
                        driver.switch_to.window(h)
                        try:
                            driver.refresh(); ensure_search_page(driver)
                            try:
                                body_before = driver.find_element(By.TAG_NAME, "body").text
                            except Exception:
                                body_before = ""
                            submit_number(driver, numer)
                            _ = wait_body_text_change(driver, body_before, timeout=BODY_CHANGE_TIMEOUT)
                            st['attempt'] += 1
                            st['start_ts'] = now
                            st['deadline'] = now + RESULT_TIMEOUT
                        except Exception:
                            st['attempt'] += 1
                            st['start_ts'] = now
                            st['deadline'] = now + RESULT_TIMEOUT
                        progress = True
                    else:
                        # Fail – timeout
                        operator = f"Błąd po {MAX_RETRIES} próbach: Timeout"
                        rows = num_to_rows.get(numer, [])
                        for r in rows: sheet.cell(r, 2).value = operator
                        try:
                            cur.execute("INSERT OR REPLACE INTO cache(numer, operator) VALUES(?,?)", (numer, operator))
                            pending_commits += 1
                            if pending_commits >= CACHE_COMMIT_EVERY:
                                conn.commit(); pending_commits = 0
                        except Exception as e:
                            logger.warning(f"Cache write fail {numer}: {e}")
                        processed_unique += 1
                        wrote_since_save += max(1, len(rows))
                        window_fail += 1
                        window_timeouts += 1
                        logger.info(f"Postęp: {processed_unique}/{total_unique} | {numer} -> {operator} (timeout)")
                        tabs[h] = {'numer': None, 'attempt': 0, 'start_ts': 0.0, 'deadline': 0.0, 'latency': None}
                        progress = True
                    continue

                # 3) Nieblokujący odczyt wyniku
                try:
                    driver.switch_to.window(h)
                    txt = find_result_text(driver)
                    operator = parse_operator(txt)
                    rows = num_to_rows.get(numer, [])
                    for r in rows: sheet.cell(r, 2).value = operator

                    try:
                        cur.execute("INSERT OR REPLACE INTO cache(numer, operator) VALUES(?,?)", (numer, operator))
                        pending_commits += 1
                        if pending_commits >= CACHE_COMMIT_EVERY:
                            conn.commit(); pending_commits = 0
                    except Exception as e:
                        logger.warning(f"Cache write fail {numer}: {e}")

                    processed_unique += 1
                    wrote_since_save += max(1, len(rows))
                    latency = time.time() - st['start_ts']
                    latencies_window.append(latency)
                    window_success += 1
                    logger.info(f"Postęp: {processed_unique}/{total_unique} | {numer} -> {operator} (lat {latency:.2f}s)")

                    tabs[h] = {'numer': None, 'attempt': 0, 'start_ts': 0.0, 'deadline': 0.0, 'latency': None}
                    progress = True

                    # Natychmiast kolejny numer – ale tylko dla aktywnych kart
                    if (h in active_set) and q:
                        try:
                            try:
                                body_before = driver.find_element(By.TAG_NAME, "body").text
                            except Exception:
                                body_before = ""
                            peek = q[0]
                            submit_number(driver, peek)
                            _ = wait_body_text_change(driver, body_before, timeout=BODY_CHANGE_TIMEOUT)
                            numer_new = q.popleft()
                            tabs[h] = {'numer': numer_new, 'attempt': 1, 'start_ts': time.time(),
                                       'deadline': time.time() + RESULT_TIMEOUT, 'latency': None}
                            time.sleep(random.uniform(*jitter_for_tabs(active_tabs)))
                        except Exception:
                            try:
                                driver.refresh(); ensure_search_page(driver)
                            except Exception: pass

                except TimeoutException:
                    # wynik jeszcze nie gotowy – normalne
                    pass
                except Exception as e:
                    # miękki błąd odczytu – spróbujemy dalej
                    logger.debug(f"Karta {h} odczyt err: {e}")

            # Zapis partii
            if wrote_since_save >= SAVE_EVERY_RESULTS:
                try: wb.save(EXCEL_OUTPUT)
                except Exception as e: logger.warning(f"Zapis pośredni nieudany: {e}")
                wrote_since_save = 0

            # AUTOTUNE co zadany wolumen wyników
            window_total = window_success + window_fail
            if window_total >= AUTOTUNE_EVERY:
                now = time.time()
                elapsed = now - last_tune_time if now > last_tune_time else 1.0
                tps = window_success / max(0.001, elapsed)
                med_lat = median_safe(list(latencies_window), default=None)
                err_rate = window_fail / max(1, window_total)

                decision = "KEEP"
                inc_ok = (err_rate <= 0.04) and (last_window_tps is None or tps >= last_window_tps * 0.98) \
                         and (last_window_median is None or med_lat is None or med_lat <= last_window_median * 1.15)
                dec_ok = (err_rate >= 0.10) or (last_window_tps is not None and tps <= last_window_tps * 0.85) \
                         or (last_window_median is not None and med_lat is not None and med_lat >= last_window_median * 1.35)

                if inc_ok and active_tabs < min(MAX_TABS, len(ALL_HANDLES)) and q:
                    active_tabs += 1
                    decision = f"INC->{active_tabs}"
                elif dec_ok and active_tabs > MIN_TABS:
                    active_tabs -= 1
                    decision = f"DEC->{active_tabs}"

                logger.info(f"AUTO: tabs={active_tabs} | tps={tps:.2f}/s, med_lat={med_lat:.2f if med_lat else float('nan')}s, err={err_rate*100:.1f}%, decision={decision}")

                # reset okna
                last_window_tps = tps
                last_window_median = med_lat
                last_tune_time = now
                window_success = window_fail = window_timeouts = 0

            if not progress:
                time.sleep(POLL_INTERVAL)

        # FINAL
        if pending_commits: conn.commit()
        try: wb.save(EXCEL_OUTPUT)
        except Exception as e:
            logger.error(f"Zapis końcowy nieudany: {e}")
            try:
                wb.save(EXCEL_OUTPUT.replace(".xlsx", "_backup.xlsx"))
                logger.info("Zapisano kopię zapasową arkusza.")
            except Exception: pass

        elapsed_total = time.time() - t0
        logger.info(f"Zakończono. Unikalnych: {processed_unique}/{total_unique}. Czas: {elapsed_total:.1f}s")

    finally:
        try: driver.quit()
        except Exception: pass
        conn.close()

if __name__ == "__main__":
    main()
