# -*- coding: utf-8 -*-
"""
Zbieranie z CRM -> Excel, z pominięciem wierszy oznaczonych przez Tampermonkey jako ' DB'.
Wersja: 2025-09-11 TM-final (bez OCR)

Działanie:
- Startuje monitor schowka (Windows GetClipboardSequenceNumber).
- Po CTRL+A -> CTRL+C w CRM: czyta tekst ze schowka.
- Pomija rekordy, przy których Tampermonkey dopiął ' DB' (na tej samej linii co NIP albo w najbliższym nagłówku bloku).
- Z pozostałych wyciąga: NIP, numer (48xxxxxxxxx -> 9 cyfr), prostą promocję + datę „do dd-mm-rrrr”, liczbę aktywnych usług (suma „X z Y”, X>0).
- Dopisuje do Excela: [NIP, Numer, Promocja, Data, Uslugi]
"""

import os, re, time, ctypes, logging, requests
from typing import List, Tuple
from openpyxl import load_workbook, Workbook
import pyperclip

# ===== ŚCIEŻKI =====
BASE_DIR   = r"C:\Users\dell\OneDrive - P1 Telecom\ZARZĄDZANIE\Programy\Generowanie baz\Czerwone i zielone kropki\Programy Wojtek\Zielone"
EXCEL_PATH = os.path.join(BASE_DIR, "zielone.xlsx")
LOG_PATH   = os.path.join(BASE_DIR, "log.txt")

os.makedirs(BASE_DIR, exist_ok=True)

# ===== LOG =====
logging.basicConfig(
    filename=LOG_PATH,
    filemode="a",
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
log = logging.getLogger(__name__)

# ===== Windows: licznik zmian schowka =====
_user32 = ctypes.windll.user32 if hasattr(ctypes, "windll") else None
def get_clipboard_seq() -> int:
    try:
        return int(_user32.GetClipboardSequenceNumber()) if _user32 else -1
    except Exception:
        return -1

# ===== Excel helpers =====
def ensure_excel(path: str):
    if not os.path.exists(path):
        wb = Workbook()
        sh = wb.active
        sh.append(["NIP", "Numer", "Promocja", "Data", "Uslugi"])
        wb.save(path)

def append_rows(rows: List[list]):
    if not rows: return
    try:
        wb = load_workbook(EXCEL_PATH)
        sh = wb.active
        for r in rows:
            sh.append(r)
        wb.save(EXCEL_PATH)
        print(f"✅ Dopisano {len(rows)} rekord(y) do Excela.")
    except Exception as e:
        log.error(f"Excel write error: {e}")
        print("❌ Błąd zapisu do Excela:", e)

# ===== GUS (opcjonalnie – jak u Ciebie wcześniej) =====
GUS_API_KEY = "bf96e683d9a9449b8958"
class GUSConnector:
    def __init__(self, api_key: str = GUS_API_KEY):
        self.api_key = api_key
        self.session = requests.Session()
        self.url = "https://wyszukiwarkaregon.stat.gov.pl/wsBIR/UslugaBIRzewnPubl.svc"
        self.session.headers.update({"Content-Type": "application/soap+xml;charset=UTF-8"})
        self.sid = None
        self._logged = False

    def _login(self) -> bool:
        env = (
            '<soap:Envelope xmlns:soap="http://www.w3.org/2003/05/soap-envelope" '
            'xmlns:ns="http://CIS/BIR/PUBL/2014/07" '
            'xmlns:wsa="http://www.w3.org/2005/08/addressing"><soap:Header>'
            '<wsa:Action>http://CIS/BIR/PUBL/2014/07/IUslugaBIRzewnPubl/Zaloguj</wsa:Action>'
            f'<wsa:To>{self.url}</wsa:To></soap:Header><soap:Body><ns:Zaloguj>'
            f'<ns:pKluczUzytkownika>{self.api_key}</ns:pKluczUzytkownika>'
            '</ns:Zaloguj></soap:Body></soap:Envelope>'
        )
        try:
            r = self.session.post(self.url, data=env, timeout=6)
            if r.status_code == 200:
                import re
                m = re.search(r"<ZalogujResult>([^<]+)</ZalogujResult>", r.text)
                if m:
                    self.sid = m.group(1)
                    self.session.headers.update({"sid": self.sid})
                    self._logged = True
                    return True
        except Exception as e:
            log.error(f"GUS login: {e}")
        return False

    def search_by_nip(self, nip: str):
        if not self._logged:
            self._login()
        env = (
            '<soap:Envelope xmlns:soap="http://www.w3.org/2003/05/soap-envelope" '
            'xmlns:ns="http://CIS/BIR/PUBL/2014/07" '
            'xmlns:wsa="http://www.w3.org/2005/08/addressing" '
            'xmlns:dat="http://CIS/BIR/PUBL/2014/07/DataContract"><soap:Header>'
            '<wsa:Action>http://CIS/BIR/PUBL/2014/07/IUslugaBIRzewnPubl/DaneSzukajPodmioty</wsa:Action>'
            f'<wsa:To>{self.url}</wsa:To></soap:Header><soap:Body><ns:DaneSzukajPodmioty>'
            '<ns:pParametryWyszukiwania>'
            f'<dat:Nip>{nip}</dat:Nip>'
            '</ns:pParametryWyszukiwania></ns:DaneSzukajPodmioty></soap:Body></soap:Envelope>'
        )
        try:
            r = self.session.post(self.url, data=env, timeout=6)
            txt = r.text if r.status_code == 200 else ""
            if "<Fault" in txt or "<faultcode>" in txt:
                if not self._login(): return None
                r = self.session.post(self.url, data=env, timeout=6)
                txt = r.text if r.status_code == 200 else ""
            if not txt: return None
            import re, html, xml.etree.ElementTree as ET
            m = re.search(r"<DaneSzukajPodmiotyResult>(.*?)</DaneSzukajPodmiotyResult>", txt, flags=re.DOTALL)
            if not m: return None
            inner = html.unescape(m.group(1)).strip()
            if not inner: return None
            root = ET.fromstring(f"<root>{inner}</root>")
            dane = root.find(".//dane")
            if dane is None: return None
            name = dane.findtext("Nazwa") or ""
            inactive = bool(dane.findtext("DataZakonczeniaDzialalnosci"))
            return {"Nazwa": name, "Nieaktywna": inactive}
        except Exception as e:
            log.error(f"GUS search: {e}")
            return None

# ===== Parsowanie z tekstu (bez DB -> dodajemy; z DB -> pomijamy) =====
NIP_RE = re.compile(r"\b\d{10}\b")

def _promo_from_details(line: str) -> Tuple[str, str]:
    m_do = re.search(r"do (\d{2}-\d{2}-\d{4})", line)
    data_do = m_do.group(1) if m_do else ""
    dates = list(re.finditer(r"\d{2}-\d{2}-\d{4}", line))
    tail = line[dates[-1].end():].strip() if dates else line.strip()
    promo = re.split(r"\s{2,}", tail)[0].strip()
    return promo, data_do

def extract_client_data(text: str, gus: GUSConnector):
    clients, skipped = [], []
    blocks = re.split(r"Wszystkich kont:\s*\d+", text, flags=re.I)
    for blk in blocks:
        blk = blk.strip()
        if not blk: 
            continue

        # NIP
        m = NIP_RE.search(blk)
        if not m:
            continue
        nip = m.group(0)

        # Czy blok jest oznaczony ' DB' przez TM?
        # 1) na tej samej linii co NIP
        nip_line = next((l for l in blk.splitlines() if nip in l), "")
        has_db_on_line = bool(re.search(rf"{nip}\b.*\bDB\b", nip_line))

        # 2) w krótkim nagłówku (do 'Numer konta')
        header_end = blk.find("Numer konta")
        header = blk if header_end == -1 else blk[:header_end]
        has_db_in_header = bool(re.search(r"\bDB\b", header))

        if has_db_on_line or has_db_in_header:
            skipped.append(f"{nip} (DB)")
            log.info(f"Skip (DB) NIP={nip}")
            continue

        # GUS: opcjonalnie odfiltruj nieaktywne
        gus_data = gus.search_by_nip(nip)
        if not gus_data or gus_data["Nieaktywna"]:
            skipped.append(nip)
            continue

        # Numer 48xxxxxxxxx -> weź preferencyjnie zaczynające się od 485
        lines = blk.splitlines()
        active_account = True
        numbers_info = []
        for idx, line in enumerate(lines):
            m_acc = re.search(r"(\d+)\s+z\s+(\d+)", line)
            if m_acc:
                active_account = m_acc.group(1) != "0"
                continue
            m_num = re.match(r"\s*48\d{9}\s*$", line)
            if m_num and active_account:
                full_num = m_num.group().strip()
                details = lines[idx + 1] if idx + 1 < len(lines) else ""
                promo, data_do = _promo_from_details(details)
                numbers_info.append((full_num, promo, data_do))

        if not numbers_info:
            skipped.append(nip)
            continue

        chosen = next((n for n in numbers_info if n[0].startswith("485")), numbers_info[0])
        numer_9 = chosen[0][2:]
        promo_val, data_do = chosen[1], chosen[2]
        uslugi = sum(int(m.group(1)) for m in re.finditer(r"(\d+)\s+z\s+(\d+)", blk) if m.group(1) != "0")

        clients.append([nip, numer_9, promo_val, data_do, uslugi])

    return clients, skipped

# ===== Monitor schowka =====
def monitor():
    ensure_excel(EXCEL_PATH)
    gus = GUSConnector()
    last_seq = get_clipboard_seq()
    print("📋 Monitor działa. W CRM: CTRL+A → CTRL+C.")

    while True:
        try:
            seq = get_clipboard_seq()
            if seq != -1 and seq != last_seq:
                last_seq = seq
                txt = ""
                try:
                    txt = pyperclip.paste() or ""
                except Exception:
                    pass

                # Sensowny zrzut ma zwykle > 100 znaków
                if len(txt.strip()) < 20:
                    print("⚠️ Nic do zrobienia (pusty/krótki schowek).")
                    time.sleep(0.2)
                    continue

                # Parsuj i zapisuj
                active_rows, skipped_nips = extract_client_data(txt, gus)
                if active_rows:
                    append_rows(active_rows)
                if skipped_nips:
                    print("⛔ Pominięto:", ", ".join(skipped_nips))
                elif not active_rows:
                    print("⚠️ Brak aktywnych rekordów do dopisania.")

            time.sleep(0.08)
        except Exception as e:
            log.exception(f"Monitor error: {e}")
            print("❌ Błąd monitora:", e)
            time.sleep(0.5)

if __name__ == "__main__":
    monitor()
